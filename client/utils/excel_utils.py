from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta

def generate_excel(data, keys, colored, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicles"

    # Ensure 'hu' and 'rnr' columns are always present
    if 'hu' not in keys:
        keys.append('hu')
    if 'rnr' not in keys:
        keys.insert(0, 'rnr')  # Ensure 'rnr' is the first column

    # Prepare header
    ws.append(keys)

    # Add rows sorted by 'gruppe'
    data_sorted = sorted(data, key=lambda x: x.get('gruppe', ''))
    today = datetime.today()

    for item in data_sorted:
        row = []
        for key in keys:
            value = item.get(key, "")
            row.append(value)
        ws.append(row)

        # Apply row coloring logic if -c flag is True
        if colored:
            hu_date = item.get("hu")
            row_color = None

            if hu_date:
                try:
                    hu_date = datetime.strptime(hu_date, "%Y-%m-%d")
                    if hu_date >= today - timedelta(days=90):  # Not older than 3 months
                        row_color = "007500"  # Green
                    elif hu_date >= today - timedelta(days=365):  # Not older than 12 months
                        row_color = "FFA500"  # Orange
                    else:  # Older than 12 months
                        row_color = "b30000"  # Red
                except ValueError:
                    pass  # If hu_date is invalid, skip coloring

            if row_color:
                fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                for cell in ws[ws.max_row]:  # Use ws.max_row to get the last row
                    cell.fill = fill

    # Save to the specified output file
    wb.save(output_file)
